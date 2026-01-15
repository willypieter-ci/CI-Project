# processor.py
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import re
import os

def process_attendance(input_csv_path, output_dir_name="hasil_absensi_per_karyawan"):
    """
    Memproses file CSV absensi dan menghasilkan file Excel per karyawan.
    
    Args:
        input_csv_path (str): Path ke file CSV input
        output_dir_name (str): Nama folder output
    
    Returns:
        str: Path folder hasil
    """
    # Baca file CSV
    # df = pd.read_csv(input_csv_path)
    df = pd.read_csv(input_csv_path, encoding='utf-8')

    # Buat folder output
    output_dir = Path(output_dir_name)
    output_dir.mkdir(exist_ok=True)

    # Hapus file lama (opsional, biar bersih)
    for f in output_dir.glob("*.xlsx"):
        try:
            os.remove(f)
        except:
            pass

    # Proses per karyawan berdasarkan Card ID
    for card_id, group in df.groupby('Card ID'):
        print(f" Memproses: {card_id}")

        group = group.reset_index(drop=True)
        group['No.'] = range(1, len(group) + 1)

        def sanitize_filename(name):
            name = str(name)
            if name.isdigit():
                name = f"Karyawan_{name}"
            name = re.sub(r'[^\w\s-]', '_', name)
            name = re.sub(r'[\s-]+', '_', name)
            return name.strip('_')
        
        safe_name = sanitize_filename(card_id)
        output_path = output_dir / f"{safe_name}_absensi_rapi.xlsx"

        # Buat workbook baru
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Header atau judul
        ws.merge_cells('A1:L1')
        ws['A1'].value = "Absensi Karyawan Kontraktor"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = center_align

        ws['A2'].value = "Operator : admin"
        ws['A3'].value = "Criteria :Tanggal 01 Desember s/d 16 Desember 2025"
        ws['A4'].value = "PT.CANANG INDAH"

        for row in [2, 3, 4]:
            ws[f'A{row}'].font = bold_font
            ws[f'A{row}'].alignment = left_align

        # --- tabel header ---
        headers = [
            "No.", "Card ID", "Employee ID", "Name", "Depart.", "Date",
            "First IN", "Last OUT", "Terminal(First)", "Terminal(Last)", "Door(First)", "Door(Last)"
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num, value=header)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border

        start_row = 6
        for idx, row in group.iterrows():
            for col_num, value in enumerate(row.values, 1):
                cell = ws.cell(row=start_row + idx, column=col_num, value=value)
                cell.alignment = center_align
                cell.border = thin_border

        # --- footer ---
        footer_row = start_row + len(group) + 1
        ws.cell(row=footer_row, column=1).value = "Total Jam Kerja :"
        ws.cell(row=footer_row + 1, column=1).value = "Total Hari Kerja :"
        ws.cell(row=footer_row + 2, column=1).value = "Total Libur & Time :"

        for row in [footer_row, footer_row + 1, footer_row + 2]:
            ws.cell(row=row, column=1).font = bold_font

        col_widths = [5, 12, 12, 20, 10, 18, 12, 12, 18, 18, 12, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Simpan file
        wb.save(output_path)
        print(f"✅ File '{output_path}' berhasil dibuat!")

    return str(output_dir)